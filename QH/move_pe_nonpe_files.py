import openpyxl
import os

wb = openpyxl.load_workbook('scan_report.xlsx')
sheet = wb['undetected']
PE_values = []
NON_PE_values = []

for row in range(2, sheet.max_row + 1):
    cell_value = sheet.cell(row=row, column=1).value
    if cell_value != None:
        PE_values.append(cell_value)

for row in range(2, sheet.max_row + 1):
    cell_value = sheet.cell(row=row, column=2).value
    if cell_value != None:
        NON_PE_values.append(cell_value)

print("PE_values:", len(PE_values))
print("NON_PE_values:", len(NON_PE_values))

for i in PE_values:
    try:
        os.rename(f"./sample/{i}", f"./PE/{i}")
    except Exception as e:
        print(e)

for i in NON_PE_values:
    try:
        os.rename(f"./sample/{i}", f"./NON_PE/{i}")
    except Exception as e:
        print(e)


print("\n\nDone...")
